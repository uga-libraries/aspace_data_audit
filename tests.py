import ast
import time
import unittest

import openpyxl
import pathlib
import subprocess

from ASpace_Data_Audit import *

AUDIT_FILE = ""  # Global variable for storing the filepath of the audit spreadsheet generated


class AuditOutputTests(unittest.TestCase):

    def test_run_report(self):
        """
        Run the ASpace_Data_Audit.py file using --test parameter and determine if it generated a spreadsheet report.
        """
        testing_spreadsheet = pathlib.Path(os.getcwd(), f'data_audit_{str(date.today())}.xlsx')
        vevn_python = pathlib.Path(os.getcwd(), 'venv/Scripts/python.exe')
        subprocess.run([vevn_python, '-m', 'pip', 'install', '-r', 'requirements.txt'], shell=True)
        subprocess.call([vevn_python, 'ASpace_Data_Audit.py', '--test'], shell=True)
        report_generated = self.assertIsFile(testing_spreadsheet)
        if report_generated is True:
            global AUDIT_FILE
            AUDIT_FILE = report_generated

    @staticmethod
    def assertIsFile(path):
        """
        Determine if the given filepath exists and if not, return an error.
        Args:
            path (Path): the Path object of the filepath being tested
        """
        if not pathlib.Path(path).resolve().is_file():
            raise AssertionError(f'File does not exist: {str(path)}')

    @staticmethod
    def assertIsFolder(path):
        """
        Determine if the given folder path exists and if not, return an error.
        Args:
            path (Path): the Path object of the folder path being tested
        """
        if not pathlib.Path(path).resolve().is_dir():
            raise AssertionError(f'Folder does not exist: {str(path)}')

    @staticmethod
    def assertHasFiles(path):
        """
        Determine if files exist in the given folder path and if not, return an error.
        Args:
            path (Path): the Path object of the folder path being tested
        """
        if not os.listdir(path):
            raise AssertionError(f'Files do not exist in {path}')


class TestASpaceFunctions(AuditOutputTests):

    def test_connect_aspace_api(self):
        """
        Test the connect_aspace_api() function and that it returns an ASnakeClient instance, using the
        credentials in the secrets.py file.
        """
        self.local_aspace = connect_aspace_api()
        self.assertIsInstance(self.local_aspace, ASnakeClient)

    def test_export_eads(self):
        """
        Create a source_eads folder if one doesn't exist and test the export_eads() function by determining if
        source_eads has files with .xml file extensions.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        export_folder = str(pathlib.Path(os.getcwd(), "source_eads"))
        if not os.path.exists(export_folder):
            os.mkdir(export_folder)
        local_aspace = connect_aspace_api()
        export_eads(test_workbook, export_folder, local_aspace)
        self.assertHasFiles(Path(export_folder))
        for root, directories, files in os.walk(export_folder):
            for filename in files:
                self.assertEqual(str(Path(filename).suffix), '.xml')

    # The following tests check the functions grabbing and testing the ArchivesSpace data

    def test_check_creators(self):
        """
        Test the check_creators() function by creating a test workbook/spreadsheet to see if results in the spreadsheet
        where collections without creators returns None value.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        self.local_aspace = connect_aspace_api()
        check_creators(test_workbook, self.local_aspace)
        test_workbook.save(test_spreadsheet_filepath)

        # test_workbook = openpyxl.load_workbook(self.test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames
        self.assertIn("Resources without Creators", test_sheetnames)

        if "Resources without Creators" in test_sheetnames:
            test_sheet = test_workbook["Resources without Creators"]
            for row in test_sheet.iter_rows(min_row=2, max_row=2, min_col=4, max_col=4):
                for cell in row:
                    if cell.value:
                        self.assertEqual(cell.value, "None")
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)

    def test_check_res_levels(self):
        """
        Tests the check_res_levels() function by creating a test workbook/spreadsheet and running the function to
        determine if results found in the generated spreadsheet are a list and that they contain equal or more than 2
        items within that list.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        self.local_aspace = connect_aspace_api()
        check_res_levels(test_workbook, self.local_aspace, test=True)
        test_workbook.save(test_spreadsheet_filepath)

        # test_workbook = openpyxl.load_workbook(self.test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames
        self.assertIn("Collection Level Checks", test_sheetnames)

        if "Collection Level Checks" in test_sheetnames:
            test_sheet = test_workbook["Collection Level Checks"]
            for row in test_sheet.iter_rows(min_row=2, max_row=2, min_col=5, max_col=5):
                for cell in row:
                    if cell.value:
                        test_value = ast.literal_eval(cell.value)
                        self.assertIsInstance(test_value, list)
                        self.assertTrue(len(cell.value) >= 2)
        os.remove(test_spreadsheet_filepath)

    # def test_check_child_levels(self):  # This is part of test_check_res_levels
    #     pass


class SpreadsheetTests(unittest.TestCase):

    def test_generate_spreadsheet(self):
        """
        Test the generate_spreadsheet() function creates a data_audit_date.xlsx file, an openpyxl workbook instance,
        and that the filepath is a string.
        """
        self.test_workbook, self.test_spreadsheet_filepath = generate_spreadsheet()
        self.assertIsInstance(self.test_workbook, openpyxl.Workbook)
        self.assertIsInstance(self.test_spreadsheet_filepath, str)
        self.assertEqual(os.path.exists(self.test_spreadsheet_filepath), True)

    def test_write_headers(self):
        """
        Test write_headers() function by writing a list of test headers to a spreadsheet and assert that those test
        header values exist in the spreadsheet.
        """
        test_headers = ["test_header_1", "test_header_2", "test_header_3"]
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        write_headers(test_workbook, "test", test_headers)
        test_workbook.save(test_spreadsheet_filepath)

        test_sheetnames = test_workbook.sheetnames
        self.assertIn("test", test_sheetnames)

        if "test" in test_sheetnames:
            test_sheet = test_workbook["test"]
            for row in test_sheet.iter_rows(max_row=1, max_col=3):
                for cell in row:
                    self.assertIn(cell.value, test_headers)
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)


class SQLTests(unittest.TestCase):

    def test_db_connection(self):
        """
        Test the connect_db() function by asserting that the db connection and cursor returned are not None.
        """
        self.db_connect, self.db_cursor = connect_db()
        self.assertIsNotNone(self.db_connect)
        self.assertIsNotNone(self.db_cursor)

    def test_query_db(self):
        """
        Test the query_database() function by supplying a test query and determining that there are results returned
        and those results are a list.
        """
        test_statement = ('SELECT name, username FROM user')
        self.db_connect, self.db_cursor = connect_db()
        results = query_database(self.db_connect, self.db_cursor, test_statement)
        self.assertIsNotNone(results)
        self.assertIsInstance(results, list)

    def test_run_query(self):
        """
        Test the run_query() function by creating a test workbook/spreadsheet, headers, and a statement checking for
        users in the ASpace database and determining if Administrator is one of the returned results in the spreadsheet.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        test_headers = ["Name", "Username", "System Administrator?", "Hidden User?"]
        self.db_connect, self.db_cursor = connect_db()
        self.test_statement = ('SELECT name, username, is_system_user AS System_Administrator, '
                               'is_hidden_user AS Hidden_User FROM user')
        run_query(test_workbook, 'Users', test_headers, self.test_statement, booleans=True)
        test_workbook.save(test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames
        if "Users" in test_sheetnames:
            user_sheet = test_workbook["Users"]
            potential_users = []
            for row in user_sheet.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    potential_users.append(cell.value)
            self.assertIn('Administrator', potential_users)
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)

    # The following tests check the functions grabbing and testing the ArchivesSpace data

    def test_check_controlled_vocabs(self):
        """
        Test check_controlled_vocabs() by creating a test workbook/spreadsheet, test terms and term number, and run
        those terms against the function, checking the spreadsheet to determine if lcnaf was returned and that "ingest"
        returns a red-highlighted row.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        test_vocab = 'Name_Sources'
        test_terms = ["local", "naf", "ingest", "snac", "lcnaf"]
        test_term_num = 4
        check_controlled_vocabs(test_workbook, test_vocab, test_terms, test_term_num)

        test_workbook.save(test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames
        if "Name_Sources" in test_sheetnames:
            user_sheet = test_workbook["Name_Sources"]
            potential_vocab = []
            for row in user_sheet.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    potential_vocab.append(cell.value)
                    if cell.value not in test_terms:
                        self.assertTrue(cell.font.color)
            self.assertIn('lcnaf', potential_vocab)
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)

    def test_check_duplicates(self):
        """
        Test check_duplicates() by creating a test workbook/spreadsheet, headers, statement, sheetname, and uri string
        to run through check_duplicates, asserting that result titles are strings and if there are duplicates, checking
        the duplicate results against each other to make sure they are the same value, but have different URIs.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        test_headers = ["Original Subject", "Original Subject ID", "Duplicate Subject", "Duplicate Subject ID"]
        test_statement = f'SELECT title, id FROM subject'
        test_sheetname = 'Duplicate Subjects'
        test_uri_string = '/subjects/'
        check_duplicates(test_workbook, test_headers, test_statement, test_sheetname, test_uri_string)

        test_workbook.save(test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames
        if 'Duplicate Subjects' in test_sheetnames:
            user_sheet = test_workbook["Duplicate Subjects"]
            potential_duplicates = {}
            duplicate_count = 0
            for row in user_sheet.iter_rows(min_row=2, max_col=4):
                potential_duplicates[duplicate_count] = []
                for cell in row:
                    if cell.value:
                        self.assertIsInstance(cell.value, str)
                        potential_duplicates[duplicate_count].append(cell.value)
                duplicate_count += 1
            for result_index, results in potential_duplicates.items():
                if results:  # if results not an empty list
                    original_name = results[0]
                    duplicate_name = results[2]
                    self.assertEqual(original_name, duplicate_name)

                    original_uri = results[1]
                    duplicate_uri = results[3]
                    self.assertNotEqual(original_uri, duplicate_uri)
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)


class AuditFunctionsTests(AuditOutputTests):

    def test_email_users(self):
        """
        Test email_users() by requiring user input for an email to send from, to, and an email server to use, asking
        the user to see if they receive the email and typing "Yes" or "No" in the console to pass or fail.
        """
        send_from = input(f'Enter the email to send from: ')
        send_to = [f'{input("Enter the email to send to: ")}']
        email_subject = f'Test email_users for ASpace_Data_Audit'
        email_message = ("This is a test of the email_users function. If you received this, you can type 'Yes' in the "
                         "console response.")
        test_server = input(f'Enter the email server: ')
        try:
            email_users(send_from, send_to, email_subject, email_message, server=test_server)
        except Exception as error:
            self.fail(error)
        email_response = input(f'Did you receive an email from the above source? It may take a minute or two. '
                               f'Type Yes or No: ').lower()
        self.assertEqual(email_response, 'yes')

    def test_email_error(self):
        """
        Test email_error() by requiring user input for an email to send from, to, and a email server to use, asking
        the user to see if they receive the email and typing "Yes" or "No" in the console to pass or fail.
        """
        send_from = input(f'Enter the email to send from: ')
        send_to = [f'{input("Enter the email to send to: ")}']
        test_server = input(f'Enter the email server: ')
        try:
            email_error(send_from, send_to, f'TEST ERROR for ASpace_Data_Audit - not real error',
                        server=test_server)
        except Exception as error:
            self.fail(error)
        email_response = input(f'Did you receive an email from the above source? It may take a minute or two. '
                               f'Type Yes or No: ').lower()
        self.assertEqual(email_response, 'yes')

    def test_standardize_resids(self):
        """
        Test standardize_resids() by creating a test SQL query to return repository name and resource identifiers,
        checking the original SQL results with the output of standardize_resids to make sure they are not equal, that
        the output of standardized_resids are a string, and that "Null" does not exist in output.
        """
        test_statement = ('SELECT repo.name AS Repository, resource.identifier AS Resource_ID  '
                          'FROM repository AS repo '
                          'JOIN resource ON repo.id = resource.repo_id ')
        self.db_connect, self.db_cursor = connect_db()
        results = query_database(self.db_connect, self.db_cursor, test_statement)
        updated_resids = standardize_resids(results)
        self.assertNotEqual(results[1][1], updated_resids[1][1])
        self.assertIsInstance(updated_resids[1][1], str)
        self.assertNotIn("Null", updated_resids[1][1])

    def test_update_booleans(self):
        """
        Test update_booleans() by creating a test SQL query and running it through query_database - taking those results
        and passing them through update_booleans, determining that the original SQL results do not match with the output
        of update_booleans and that the output type is boolean.
        """
        test_statement = ('SELECT name, username, is_system_user AS System_Administrator '
                          'FROM user')
        self.db_connect, self.db_cursor = connect_db()
        results = query_database(self.db_connect, self.db_cursor, test_statement)
        updated_booleans = update_booleans(results)
        self.assertIsNot(results[1][2], updated_booleans[1][2])
        self.assertIsInstance(updated_booleans[1][2], bool)

    def test_check_export_folder(self):
        """
        Test create_export_folder() by running the function and assert if a source_eads folder now exists in the
        current directory.
        """
        create_export_folder()
        export_folder = pathlib.Path(os.getcwd(), "source_eads")
        self.assertIsFolder(export_folder)
        delete_export_folder(str(export_folder))

    def test_delete_export_folder(self):
        """
        Test delete_export_folder() by creating a test_source_eads folder in the current directory, sleeping for 5
        seconds, then running the function, checking if no path exists for the test folder.
        """
        source_eads_path = str(Path.joinpath(Path.cwd(), "test_source_eads"))
        os.mkdir(source_eads_path)
        time.sleep(5)
        delete_export_folder(source_eads_path)
        if not os.path.exists(source_eads_path):
            pass
        else:
            self.fail()

    def test_run_audit(self):
        """
        Test running the entire audit with test_run_report() unittest
        """
        AuditOutputTests.test_run_report(self)

    def test_run_script(self):
        """
        Test run_script() by running the function and checking to see if the appropriate data_audit spreadsheet exists.
        """
        generated_report = pathlib.Path(os.getcwd(), f'data_audit_{str(date.today())}.xlsx')
        run_script(test=True)
        self.assertIsFile(generated_report)

    # The following tests check the functions grabbing and testing the ArchivesSpace data

    def test_duplicate_subjects(self):
        """
        Test duplicate_subjects() by creating a test workbook/spreadsheet and determining if there are any results in
        the spreadsheet and if so, asserting that result subjects are strings and if there are duplicates, checking
        the duplicate results against each other to make sure they are the same value, but have different URIs.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        duplicate_subjects(test_workbook)

        test_workbook.save(test_spreadsheet_filepath)
        # test_workbook = openpyxl.load_workbook(test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames
        if 'Duplicate Subjects' in test_sheetnames:
            user_sheet = test_workbook["Duplicate Subjects"]
            potential_duplicates = {}
            duplicate_count = 0
            for row in user_sheet.iter_rows(min_row=2, max_col=4):
                potential_duplicates[duplicate_count] = []
                for cell in row:
                    if cell.value:
                        self.assertIsInstance(cell.value, str)
                        potential_duplicates[duplicate_count].append(cell.value)
                duplicate_count += 1
            for result_index, results in potential_duplicates.items():
                if results:  # if results not an empty list
                    original_name = results[0]
                    duplicate_name = results[2]
                    self.assertEqual(original_name, duplicate_name)

                    original_uri = results[1]
                    duplicate_uri = results[3]
                    self.assertNotEqual(original_uri, duplicate_uri)
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)

    def test_duplicate_agent_persons(self):
        """
        Test duplicate_agent_persons() by creating a test workbook/spreadsheet and determining if there are any results
        in the spreadsheet and if so, asserting that result names are strings and if there are duplicates, checking
        the duplicate results against each other to make sure they are the same value, but have different URIs.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        duplicate_agent_persons(test_workbook)

        test_workbook.save(test_spreadsheet_filepath)
        # test_workbook = openpyxl.load_workbook(test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames
        if 'Duplicate Agents' in test_sheetnames:
            user_sheet = test_workbook["Duplicate Agents"]
            potential_duplicates = {}
            duplicate_count = 0
            for row in user_sheet.iter_rows(min_row=2, max_col=4):
                potential_duplicates[duplicate_count] = []
                for cell in row:
                    if cell.value:
                        self.assertIsInstance(cell.value, str)
                        potential_duplicates[duplicate_count].append(cell.value)
                duplicate_count += 1
            for result_index, results in potential_duplicates.items():
                if results:  # if results not an empty list
                    original_name = results[0]
                    duplicate_name = results[2]
                    self.assertEqual(original_name, duplicate_name)

                    original_uri = results[1]
                    duplicate_uri = results[3]
                    self.assertNotEqual(original_uri, duplicate_uri)
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)

    def test_check_urls(self):
        """
        Test check_urls() by creating a test workbook/spreadsheet, export folder, exporting EADS and running the
        function, asserting that any results in the spreadsheet are checked with response.get and do not return a 200
        status code
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        export_folder = str(pathlib.Path(os.getcwd(), "source_eads"))
        if not os.path.exists(export_folder):
            os.mkdir(export_folder)
        local_aspace = connect_aspace_api()
        export_eads(test_workbook, export_folder, local_aspace)
        check_urls(test_workbook, export_folder)

        test_workbook.save(test_spreadsheet_filepath)
        # test_workbook = openpyxl.load_workbook(test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames
        if 'URL Errors' in test_sheetnames:
            user_sheet = test_workbook["URL Errors"]
            for url in user_sheet.iter_rows(min_row=2, max_row=20, min_col=4, max_col=4, values_only=True):
                try:
                    response = requests.get(url[0], allow_redirects=True, timeout=30)
                except:
                    pass
                else:
                    print(url[0])
                    self.assertNotEqual(response.status_code, 200)
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)

    def test_check_url(self):
        """
        Test check_url() by supplying a good and bad url to test, asserting the good URL returns None and a bad URL
        returns a result.
        """
        test_good_url = 'https://www.libs.uga.edu/'
        test_bad_url = 'http://www.cviog.uga.edu/about/chapel/history.php'
        good_response_code = check_url(test_good_url)
        bad_response_code = check_url(test_bad_url)
        self.assertIsNone(good_response_code)
        self.assertIsNotNone(bad_response_code)

    def test_search_ghost_containers(self):
        """
        Test search_ghost_containers() by creating a test workbook/spreadsheet and running the function, then taking
        any results and searching for them in ArchivesSpace, assert if collection field is not in returned ArchivesSpace
        data and if it is, then it's length must equal 0 to signal it's not linked to any collections.
        """
        test_workbook, test_spreadsheet_filepath = generate_spreadsheet()
        local_aspace = connect_aspace_api()
        search_ghost_containers(test_workbook, local_aspace)

        test_workbook.save(test_spreadsheet_filepath)
        # test_workbook = openpyxl.load_workbook(test_spreadsheet_filepath)
        test_sheetnames = test_workbook.sheetnames

        if 'Unlinked Top Containers' in test_sheetnames:
            user_sheet = test_workbook["Unlinked Top Containers"]
            for container_uri in user_sheet.iter_rows(min_row=2, max_row=20, min_col=4, max_col=4, values_only=True):
                if container_uri[0]:
                    container_data = local_aspace.get(f'{container_uri[0]}',
                                                      params={"resolve[]": True}).json()
                    if "collection" not in container_data:
                        self.fail()
                    else:
                        self.assertTrue(len(container_data["collection"]) == 0)
        test_workbook.close()
        os.remove(test_spreadsheet_filepath)

    # def test_get_top_children(self):  # This is part of test_check_res_levels
    #     pass


if __name__ == '__main__':
    unittest.main()
